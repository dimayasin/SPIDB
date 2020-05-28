from __future__ import unicode_literals
from django.shortcuts import render, redirect, HttpResponse
from django.http import FileResponse
from django import forms
from django.db import connection
import django_excel as excel
import pyodbc
import openpyxl
from django.contrib import messages
from openpyxl import load_workbook, workbook
from openpyxl.styles import Font, Fill
# from django.forms.utils import ValidationError
from .models import spiInv, AFHS, Airlines, avref, ILSQH, SatairList
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from openpyxl.styles import Alignment
import xlwt

from datetime import date
import datetime


# from lib import bcrypt
# import re

# EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9.+_-]+@[a-zA-Z0-9._-]+\.[a-zA-Z]+$')
# PASSWORD_REGEX = re.compile(r'\d.*[A-Z]|[A-Z].*\d')

secret_key = 'TARDIS'

cursor = connection.cursor()


def index(request):
    # request.session['id']=""

    return render(request, 'index.html')


def summ(request):

    return render(request, 'summ.html')

def all_data(request):
    
    avref_part = list()
    afhs_part=list()
    airlines_part=list()
    SatairList_part=list()
    ilsqhList_part=list()
    messages=list()
    pn_list=request.POST['partnumber'].split('#')
    sourceList=request.POST.getlist('check')
    FileName=request.POST['fileName']

    today = datetime.date.today()
    messages=list()
    y=-1
    if request.POST['timerange'] == 'y1':
        y=-1
    elif  request.POST['timerange'] == 'y2':
        y=-2
    elif  request.POST['timerange'] == 'y3':
        y=-3
    elif  request.POST['timerange'] == 'y4':
        y=-4
    elif  request.POST['timerange'] == 'y5':
        y=-5

    for source in sourceList:
        if source =='list':
                for item in pn_list:
                    # SatairList_part.append(SatairPNQuery(item))
                    # dataset= SatairList.object.raw('select * from spiapps_SatairList where spiapps_SatairList.PN =%s  AND year(spiapps_SatairList.date) >= year(DATEADD(year,%s,getdate()))', [item,y])
                    dataset=SatairList.object.filter(PN=item)
                    if dataset:
                        for val in dataset:
                            if int(val.date.strftime('%Y')) >= today.year +y:
                                SatairList_part.append(val)
                    else:
                        messages.append('Part Number: ' + item+" Is Not Found In The Satair List")
        elif source =='avref':
                for item in pn_list:
                    # avref_part.append(avrefPNQuery(item))
                    dataset=avref.object.filter(PN=item)
                    if dataset:
                        for val in dataset:
                            if int(val.p_update.strftime('%Y')) >= today.year +y:
                                avref_part.append(val)
                    else:
                        messages.append('Part Number: ' + item+" Is Not Found In The avref List")
        
        elif source =='Airlines':
                for item in pn_list:
                    # airlines_part.append(AirlinesPNQuery(item))
                    # dataset= Airlines.object.raw('select * from spiapps_Airlines where spiapps_Airlines.PN like %s  AND year(spiapps_Airlines.date) >= year(DATEADD(year,%s,getdate()))', [item,y])
                    dataset=Airlines.object.filter(PN=item)
                    if dataset:
                        for val in dataset:
                            if int(val.date.strftime('%Y')) >= today.year +y:
                                airlines_part.append(val)
                    else:
                        messages.append('Part Number: ' + item+" Is Not Found In The Airlines List")
        elif source =='afhs':
                for item in pn_list:
                    # afhs_part.append(afhsPNQuery(item))
                    # dataset= AFHS.object.raw('select * from spiapps_afhs where spiapps_afhs.PN =%s  AND year(spiapps_afhs.date) >= year(DATEADD(year,%s,getdate()))', [item,y])
                    dataset=AFHS.object.filter(PN=item)
                    if dataset:
                        for val in dataset:
                            if int(val.date.strftime('%Y')) >= today.year +y:
                                afhs_part.append(val)
                    else:
                        messages.append('Part Number: ' + item+" Is Not Found In The AFHS List")

        elif source =='ils':
                for item in pn_list:
                    # sqlString='select * from spiapps_ilsqh where spiapps_ilsqh.PN= '+item+'  AND year(spiapps_ilsqh.date) = year(DATEADD(year,-1,getdate()))'
                    # dataset= ILSQH.object.raw('select * from spiapps_ilsqh where spiapps_ilsqh.PN =%s  AND year(spiapps_ilsqh.date) >= year(DATEADD(year,%s,getdate()))', [item,y])
                    dataset=ILSQH.object.filter(PN=item)
                    if dataset:
                        for val in dataset:
                            if int(val.date.strftime('%Y')) >= today.year +y:
                                ilsqhList_part.append(val)
                    else:
                        messages.append('Part Number: ' + item+" Is Not Found In The ILSQH List")

    

    context = {
        'avref_part': avref_part,
        'afhs_part':afhs_part,
        'airlines_part':airlines_part,
        'SatairList_part':SatairList_part,
        'messages':messages,
        'ilsqhList_part':ilsqhList_part,
        'FileName':FileName,
    }
    
    if request.POST['output'] == 'excel':
        SaveExcel(context)
    # elif request.POST['output'] == 'pdf':
    #     SavePDF(context)

    return render(request, 'displaydata.html', context=context)


def SavePDF(mydict):
    fileName='SPI_Data'+str(date.today().strftime('%Y-%m-%d')) +'.pdf'
    # Title='Data Report Generated on '+str(date.today().strftime('%Y-%m-%d'))
    # documentTitle='Data Report'
    

    # pdf=canvas.Canvas(fileName)
    # pdf.setTitle(documentTitle)

    pdf=SimpleDocTemplate(
    fileName,
    pagesize=letter
    )
    style=TableStyle(
        [('BACKGROUND', (0,0),(12,0), colors.grey),
        ('TEXTCOLOR', (0,0),(-1,0),colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1),(-1,-1), colors.lightgrey),
        ]
    )
 
    data=list()
    title=list()
    title.append('Source')
    title.append('Date')
    title.append('PN')
    title.append('Condition')
    title.append('Part Type')
    title.append('Description')
    title.append('ATA')
    title.append('Cost')
    title.append('Fleet')
    title.append('Effectivity')
    title.append('Serialized')
    title.append('LLP')
    title.append('UOM')
    data.append(title)
    data.append(messages)

    if mydict['avref_part']:
        for item in mydict['avref_part']:
            myList=['avref',item.p_update,item.PN,item.p_condit,item.p_Type,item.p_descript,'',item.Price,' ',' ',' ',' ',item.p_unit]
            data.append(myList)
    if mydict['afhs_part']:
        for item in mydict['afhs_part']:
            myList=[item.source,item.date,item.PN,' ',item.part_type,item.Description,item.ata,item.cost,item.fleet,' ',' ',' ',item.uom]
            data.append(myList)
    if mydict['airlines_part']:
        for item in mydict['airlines_part']:
            myList=[item.source,item.date,item.PN,' ',item.part_type,item.Description,item.ata,item.cost,item.fleet,' ',' ',' ',item.uom ]
            data.append(myList)
    if mydict['SatairList_part']:
        for item in mydict['SatairList_part']:
            myList=[item.source,item.date,item.PN,' ',item.part_type,item.Description,item.ata,item.Price,item.fleet,' ',' ',' ',item.uom]
            data.append(myList)
    if mydict['ilsqhList_part']:
        for item in mydict['ilsqhList_part']:
            myList=['ILSQH',item.date,item.PN,item.Condition,' ',item.Quote_Description,' ',item.Quote_Price,' ',' ',' ',' ',item.UM]
            data.append(myList)
    # if mydict['messages']:
    #     for item in mydict['messages']:
    #         data.append(item)

    table=Table(data)
    table.setStyle(style)

       # Alternate background color
    rowNumber=len(data)
    for i in range(1, rowNumber):
        if i%2 == 0:
            bc=colors.whitesmoke
        else:
            bc=colors.lightgrey
        tc=TableStyle(
            [
                ('BACKGROUND', (0,i),(-1,i), bc),

            ]
        )
    table.setStyle(tc)
    elems=[]
    elems.append(table)
    pdf.build(elems)

    # pdf.save()




def pn(request):

    return render(request, 'lookuppn.html')


def desc(request):

    return render(request, 'lookupdesc.html')


def bulk(request):

    return render(request, 'lookupbulk.html')


def inputData(request):
    return render(request, 'input_data.html')


def NewPart(request):
    errors = spiInv.object.validatePartsData(request.POST)
    if len(errors) > 0:
        for error in errors:
            messages.error(request, error)
        return render(request, "input_data.html")
    else:
        spiInv.object.create(
            source=request.POST['part_source'],
            date=datetime.datetime.strptime(
                request.POST['part_date'], '%Y-%m-%d').date(),
            PN=request.POST['part_pn'],
            cond=request.POST['part_cond'],
            part_type=request.POST['part_type'],
            Description=request.POST['part_desc'],
            ata=request.POST['part_ata'],
            cost=request.POST['part_price'],
            fleet=request.POST['part_fleet'],
            Effectivity=request.POST['part_eff'],
            Serialized=request.POST['part_serial'],
            LLP=request.POST['part_llp'],
            uom=request.POST['uom'],
        )
        part = list()

        part.append(spiInv.object.get(PN=request.POST['part_pn']))

        context = {
            "Part": part,
        }

        return render(request, 'showdata.html', context=context)


def brows(request):

    return render(request, 'brows.html')



excel_data = list()

# Bulk PN look up first part -- reading data from an excel file


def uploadData(request):
    # param =list()
    # part_dict_list={'PN':'','desc':''}
    # dict_list=[]
    # parts = list()


    if request.method == "POST":
        avref_part = list()
        afhs_part=list()
        airlines_part=list()
        SatairList_part=list()
        ilsqhList_part=list()
        messages=list()
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb['Sheet1']
        # row_data = list()
        # i = 2
        # parts_list = list()
        # messages="Recrod is Not Found"
        for row in worksheet.iter_rows():

            partNum = row[0].value
            # avref data
            dataSet=avref.object.filter(PN=partNum)        
            if dataSet == None:
                strmsg='Part number: '+partNum +' is not in avref'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    avref_part.append(val)

            # AFHS data
            dataSet=AFHS.object.filter(PN=partNum)
            if dataSet == None:
                strmsg='Part number: '+partNum +' is not in AFHS'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    afhs_part.append(val)                    
 

            dataSet=Airlines.object.filter(PN=partNum)
            if dataSet == None:
                strmsg='Part number: '+partNum +' is not in Airlines'
                messages.append(strmsg)
            else:
                for val in dataSet:        
                    airlines_part.append(val)

            dataSet=SatairList.object.filter(PN=partNum)
            if dataSet == None:
                strmsg='Part number: '+partNum +' is not in Satair'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    SatairList_part.append(val)

            dataSet=ILSQH.object.filter(PN=partNum)
            if dataSet == None:
                strmsg='Part number: '+partNum +' is not in Satair'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    ilsqhList_part.append(val)
        FileName='ExcelData'+str(date.today().strftime('%Y-%m-%d'))

        context = {
            'avref_part': avref_part,
            'afhs_part':afhs_part,
            'airlines_part':airlines_part,
            'SatairList_part':SatairList_part,
            'messages':messages,
            'FileName':FileName,
            'ilsqhList_part':ilsqhList_part,
        }
        SaveExcel(context)

        # return redirect('/bulksearch', context)
        return render(request, 'displaydata.html', context=context)
    else:
        return render(request, 'lookupbulk.html', {})

# Bulk PN lookup second part -- saving data into an excel file


def SaveExcel(mydict):
 

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title="Saved_Query"
    Cell=ws['A1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Source'
    Cell=ws['B1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Date'
    Cell=ws['C1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='PN'
    Cell=ws['D1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Condition'
    Cell=ws['E1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Part Type'
    Cell=ws['F1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Description'
    Cell=ws['G1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='ATA'
    Cell=ws['H1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Cost'
    Cell=ws['I1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Fleet'
    Cell=ws['J1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Effectivity'
    Cell=ws['K1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='Serialized'
    Cell=ws['L1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='LLP'
    Cell=ws['M1']
    Cell.font=Font(bold=True, size=14, italic=True)
    Cell.value='UOM'

    rowNum = 2

    if mydict['avref_part']:
        for item in mydict['avref_part']:

            Cell=ws['A'+str(rowNum)]
            Cell.value='avref'
            Cell=ws['B'+str(rowNum)]
            Cell.value=item.p_update
            Cell=ws['C'+str(rowNum)]
            Cell.value=item.PN
            Cell=ws['D'+str(rowNum)]
            Cell.value=item.p_condit
            Cell=ws['E'+str(rowNum)]
            Cell.value=item.p_Type
            Cell=ws['F'+str(rowNum)]
            Cell.value=item.p_descript
            Cell=ws['G'+str(rowNum)]
            Cell.value=''
            Cell=ws['H'+str(rowNum)]
            Cell.value=item.Price
            Cell=ws['I'+str(rowNum)]
            Cell.value=' '
            Cell=ws['J'+str(rowNum)]
            Cell.value=' '
            Cell=ws['K'+str(rowNum)]
            Cell.value=' '
            Cell=ws['L'+str(rowNum)]
            Cell.value=' '
            Cell=ws['M'+str(rowNum)]
            Cell.value=item.p_unit

            rowNum +=1
    if mydict['afhs_part']:
        for item in mydict['afhs_part']:

            Cell=ws['A'+str(rowNum)]
            Cell.value= item.source
            Cell=ws['B'+str(rowNum)]
            Cell.value=item.date
            Cell=ws['C'+str(rowNum)]
            Cell.value=item.PN
            Cell=ws['D'+str(rowNum)]
            Cell.value=' '
            Cell=ws['E'+str(rowNum)]
            Cell.value=item.part_type
            Cell=ws['F'+str(rowNum)]
            Cell.value=item.Description
            Cell=ws['G'+str(rowNum)]
            Cell.value=item.ata
            Cell=ws['H'+str(rowNum)]
            Cell.value=item.cost
            Cell=ws['I'+str(rowNum)]
            Cell.value=item.fleet
            Cell=ws['J'+str(rowNum)]
            Cell.value=' '
            Cell=ws['K'+str(rowNum)]
            Cell.value=' '
            Cell=ws['L'+str(rowNum)]
            Cell.value=' '
            Cell=ws['M'+str(rowNum)]
            Cell.value=item.uom
            rowNum +=1
    if mydict['airlines_part']:
        for item in mydict['airlines_part']:

            Cell=ws['A'+str(rowNum)]
            Cell.value= item.source
            Cell=ws['B'+str(rowNum)]
            Cell.value=item.date
            Cell=ws['C'+str(rowNum)]
            Cell.value=item.PN
            Cell=ws['D'+str(rowNum)]
            Cell.value=' '
            Cell=ws['E'+str(rowNum)]
            Cell.value=item.part_type
            Cell=ws['F'+str(rowNum)]
            Cell.value=item.Description
            Cell=ws['G'+str(rowNum)]
            Cell.value=item.ata
            Cell=ws['H'+str(rowNum)]
            Cell.value=item.cost
            Cell=ws['I'+str(rowNum)]
            Cell.value=item.fleet
            Cell=ws['J'+str(rowNum)]
            Cell.value=' '
            Cell=ws['K'+str(rowNum)]
            Cell.value=' '
            Cell=ws['L'+str(rowNum)]
            Cell.value=' '
            Cell=ws['M'+str(rowNum)]
            Cell.value=item.uom

            rowNum +=1
    if mydict['SatairList_part']:
        for item in mydict['SatairList_part']:

            Cell=ws['A'+str(rowNum)]
            Cell.value= item.source
            Cell=ws['B'+str(rowNum)]
            Cell.value=item.date
            Cell=ws['C'+str(rowNum)]
            Cell.value=item.PN
            Cell=ws['D'+str(rowNum)]
            Cell.value=' '
            Cell=ws['E'+str(rowNum)]
            Cell.value=item.part_type
            Cell=ws['F'+str(rowNum)]
            Cell.value=item.Description
            Cell=ws['G'+str(rowNum)]
            Cell.value=item.ata
            Cell=ws['H'+str(rowNum)]
            Cell.value=item.Price
            Cell=ws['I'+str(rowNum)]
            Cell.value=item.fleet
            Cell=ws['J'+str(rowNum)]
            Cell.value=' '
            Cell=ws['K'+str(rowNum)]
            Cell.value=' '
            Cell=ws['L'+str(rowNum)]
            Cell.value=' '
            Cell=ws['M'+str(rowNum)]
            Cell.value=item.uom

            rowNum +=1

    if mydict['ilsqhList_part']:
        for item in mydict['ilsqhList_part']:

            Cell=ws['A'+str(rowNum)]
            Cell.value='ILSQH'
            Cell=ws['B'+str(rowNum)]
            Cell.value=item.date
            Cell=ws['C'+str(rowNum)]
            Cell.value=item.PN
            Cell=ws['D'+str(rowNum)]
            Cell.value=item.Condition
            Cell=ws['E'+str(rowNum)]
            Cell.value=' '
            Cell=ws['F'+str(rowNum)]
            Cell.value=item.Quote_Description
            Cell=ws['G'+str(rowNum)]
            Cell.value=' '
            Cell=ws['H'+str(rowNum)]
            Cell.value=item.Quote_Price
            Cell=ws['I'+str(rowNum)]
            Cell.value=' '
            Cell=ws['J'+str(rowNum)]
            Cell.value=' '
            Cell=ws['K'+str(rowNum)]
            Cell.value=' '
            Cell=ws['L'+str(rowNum)]
            Cell.value=' '
            Cell=ws['M'+str(rowNum)]
            Cell.value=item.UM

            rowNum +=1
            # merge_format = wb.add_format({'bold': 1,'align': 'center','valign': 'vcenter','fg_color': 'yellow'})
        if mydict['messages']:
            for msg in mydict['messages']:
                ws.merge_cells(start_row=rowNum, start_column=1, end_row=rowNum, end_column=13)
                # Cell.value=msg
                Cell=ws['A'+str(rowNum)]
                Cell.value=msg
                Cell.alignment =Alignment(horizontal='center', vertical='center') 
                rowNum +=1

    # response = HttpResponse()#mimetype="application/ms-excel")
    # response['Content-Disposition'] = 'attachment; filename=mydict["FileName"]+".xlsx"'
    # 'SPI_Data'+str(date.today().strftime('%Y-%m-%d'))
    wb.save(filename=mydict["FileName"]+".xlsx")#response)
    wb.close()
    # return (response)
    # str='temp.xlsx'
    # show(str)
    # return render(request, 'showdata.html')


def spiinv_show_all(request):
    part=list()

    dataset=spiInv.object.filter()

    if dataset:
        for val in dataset:
            part.append(val)

    context = {
        "Part": part,
    }
    return render(request, 'showdata.html', context=context)
    

def pn_search(request):
 
    avref_part = list()
    afhs_part=list()
    airlines_part=list()
    SatairList_part=list()
    ilsqhList_part=list()
    messages=list()
    pn_list=request.POST['part_num'].split("#")
    for item in pn_list:

        dataSet=avref.object.filter(PN=item)        
        if dataSet == None:
            strmsg='Part number: '+item +' is not in avref'
            messages.append(strmsg)
        else:
            for val in dataSet:
                avref_part.append(val) #avref_part)


        dataSet=AFHS.object.filter(PN=item)
        if dataSet == None:
            strmsg='Part number: '+item +' is not in AFHS'
            messages.append(strmsg)
        else:
            for val in dataSet:
                afhs_part.append(val)

        dataSet=Airlines.object.filter(PN=item)
        if dataSet == None:
            strmsg='Part number: '+item +' is not in Airlines'
            messages.append(strmsg)
        else:
            for val in dataSet:        
                airlines_part.append(val)

        dataSet=SatairList.object.filter(PN=item)
        if dataSet == None:
            strmsg='Part number: '+item +' is not in Satair'
            messages.append(strmsg)
        else:
            for val in dataSet:
                SatairList_part.append(val)

        dataSet=ILSQH.object.filter(PN=item)
        if dataSet == None:
            strmsg='Part number: '+item +' is not in Satair'
            messages.append(strmsg)
        else:
            for val in dataSet:
                ilsqhList_part.append(val)

    context = {
        'avref_part': avref_part,
        'afhs_part':afhs_part,
        'airlines_part':airlines_part,
        'SatairList_part':SatairList_part,
        'messages':messages,
        'ilsqhList_part':ilsqhList_part,
    }


    return render(request, 'displaydata.html', context=context)
    
def desc_search(request):
 
    avref_part = list()
    afhs_part=list()
    airlines_part=list()
    SatairList_part=list()
    ilsqhList_part=list()
    # datalist=list()
    
    messages=list()
    if request.POST['part_desc']:
        pn_list=request.POST['part_desc'].split('#')
      
        for item in pn_list:

            dataSet=avref.object.filter(p_descript__contains=item)        
            if dataSet == None:
                strmsg='Part: '+item +' is not in avref'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    print('*****************')
                    print(val)
                    print('*****************')
                    avref_part.append(val) #avref_part)


            dataSet=AFHS.object.filter(Description__contains=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in AFHS'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    afhs_part.append(val)

            dataSet=Airlines.object.filter(Description__contains=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in Airlines'
                messages.append(strmsg)
            else:  
                for val in dataSet:      
                    airlines_part.append(val)

            dataSet=SatairList.object.filter(Description__contains=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in Satair'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    SatairList_part.append(val)

            dataSet=ILSQH.object.filter(Quote_Description__contains=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in Satair'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    ilsqhList_part.append(val)
        
    if request.POST['full_desc']:
        pn_list=request.POST['full_desc'].split('#')
        # if len(pn_list) == 1:
        #     print(pn_list)
        # else:
        #     print("long")
        
        # # print(len(pn_list))
        
        for item in pn_list:
            print(item)

            dataSet=avref.object.filter(p_descript=item)        
            if dataSet == None:
                strmsg='Part: '+item +' is not in avref'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    print('*****************')
                    print(val)
                    print('*****************')
                    avref_part.append(val) #avref_part)


            dataSet=AFHS.object.filter(Description=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in AFHS'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    afhs_part.append(val)

            dataSet=Airlines.object.filter(Description=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in Airlines'
                messages.append(strmsg)
            else:  
                for val in dataSet:      
                    airlines_part.append(val)

            dataSet=SatairList.object.filter(Description=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in Satair'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    SatairList_part.append(val)

            dataSet=ILSQH.object.filter(Quote_Description=item)
            if dataSet == None:
                strmsg='Part: '+item +' is not in Satair'
                messages.append(strmsg)
            else:
                for val in dataSet:
                    ilsqhList_part.append(val)

    context = {
        'avref_part': avref_part,
        'afhs_part':afhs_part,
        'airlines_part':airlines_part,
        'SatairList_part':SatairList_part,
        'messages':messages,
        'ilsqhList_part':ilsqhList_part,
    }


    return render(request, 'displaydata.html', context=context)

def show(request, PN_List):

    part = list()
    if len(PN_List) == 1:

        dataSet = spiInv.object.get(PN=PN_List[0])
        for val in dataSet:
            part.append(val)
    elif len(PN_List) > 1:
        for i in PN_List:
            row = list()
            dataSet = spiInv.object.filter(PN=i)
            for val in dataSet:
                row.append(val)
            part.append(row)
    context = {
        "Part": part,
    }

    # context={'excel_data': excel_data}
    return render(request, 'showdata.html', context=context)


def edit(request, pn):
    part = spiInv.object.get(PN=pn)
    context = {
        "Part": part,
    }
    return render(request, "edit_data.html", context)


def editPart(request, pn):
    myPart = spiInv.object.filter(id=pn)
    edited_part = myPart[0]
    if not request.POST['part_source'] == edited_part.source:
        edited_part.source = request.POST['part_source']
        edited_part.save()
    if not datetime.datetime.strptime(request.POST['part_date'], '%Y-%m-%d').date() == edited_part.date:
        edited_part.date = datetime.datetime.strptime(
            request.POST['part_date'], '%Y-%m-%d').date()
        edited_part.save()
    if not request.POST['part_pn'] == edited_part.PN:
        edited_part.PN = request.POST['part_pn']
        edited_part.save()
    if not request.POST['part_cond'] == edited_part.cond:
        edited_part.cond = request.POST['part_cond']
        edited_part.save()
    if not request.POST['part_type'] == edited_part.part_type:
        edited_part.part_type = request.POST['part_type']
        edited_part.save()
    if not request.POST['part_desc'] == edited_part.Description:
        edited_part.Description = request.POST['part_desc']
        edited_part.save()
    if not request.POST['part_ata'] == edited_part.ata:
        edited_part.ata = request.POST['part_ata']
        edited_part.save()
    if not request.POST['part_price'] == edited_part.cost:
        edited_part.cost = request.POST['part_price']
        edited_part.save()
    if not request.POST['part_fleet'] == edited_part.fleet:
        edited_part.fleet = request.POST['part_fleet']
        edited_part.save()
    if not request.POST['part_eff'] == edited_part.Effectivity:
        edited_part.Effectivity = request.POST['part_eff']
        edited_part.save()
    if not request.POST['part_serial'] == edited_part.Serialized:
        edited_part.Serialized = request.POST['part_serial']
        edited_part.save()
    if not request.POST['part_llp'] == edited_part.LLP:
        edited_part.LLP = request.POST['part_llp']
        edited_part.save()
    if not request.POST['uom'] == edited_part.uom:
        edited_part.uom = request.POST['uom']
        edited_part.save()

    part = list()
    part.append(spiInv.object.get(id=pn)) 
    context = {
        "Part": part,
    }

    return render(request, 'showdata.html', context)


def delete(request, pn):
    spiInv.object.filter(PN=pn).delete()
    messages = "Record Has Been Deleted."
    context = {
        'messages': messages,
    }
    return render(request, 'showdata.html', context)
